import os
import argparse
import pathlib
import random
import time

import torch
import pdb
import torch.nn as nn
from torch.backends import cudnn
from openpyxl import Workbook, load_workbook

from networks.TMANet.TMANet import *
from networks.TMANet.embed_utils import get_embeddings
from utils.test_3d_patch import *


os.environ['CUDA_VISIBLE_DEVICES'] = '0'

# --- Excel 文件配置 ---
EXCEL_FILENAME = os.path.join('./model/LA', "test_metrics_summary.xlsx")
EXCEL_HEADERS = ['exp_name', 'dice', 'jc', 'hd', 'asd']

# --- 测试路径列表 ---
TEST_PATHS = [
    "./model/KiTS19/KiTS19_tmanet_lab10/self_train",
    "./model/KiTS19/KiTS19_tmanet_lab20/self_train",
]

num_classes = 2

with open('./Datasets/kits/test.txt', 'r') as f:
    image_list = f.readlines()
image_list = [r"/data1/mengqingxu/Dataset/KiTS19/KiTS19-SSL/" + item.replace('\n', '') + ".h5" for item in image_list]


def create_TMAnet(ema=False):
    net = TMA_Net(
        patch_size=(128, 128, 64),
        n_channels=1,
        n_classes=2,
        normalization='instancenorm',
        has_dropout=False,
        text_fuse_level=1
    )
    net = nn.DataParallel(net)
    model = net.cuda()
    if ema:
        for param in model.parameters():
            param.detach_()
    return model


num_classes_dfu = 2


def init_excel_file():
    """检查Excel文件是否存在，如果不存在则创建并写入表头。"""
    if not os.path.exists(EXCEL_FILENAME):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Metrics"
        sheet.append(EXCEL_HEADERS)
        workbook.save(EXCEL_FILENAME)
        print(f"Created new Excel file: {EXCEL_FILENAME}")


def append_to_excel(data_dict):
    """将一行数据追加到Excel文件中。"""
    try:
        workbook = load_workbook(EXCEL_FILENAME)
        sheet = workbook.active
        row_to_append = [data_dict.get(header, "") for header in EXCEL_HEADERS]
        sheet.append(row_to_append)
        workbook.save(EXCEL_FILENAME)
    except FileNotFoundError:
        print(f"Error: Excel file {EXCEL_FILENAME} not found. Please initialize it first.")
    except Exception as e:
        print(f"An error occurred while writing to Excel: {e}")


def test_one_path(test_path):
    """测试单个路径的模型。"""
    exp_name = pathlib.Path(test_path).parent.name
    print(f"\n{'=' * 60}")
    print(f"Testing: {exp_name}")
    print(f"Test path: {test_path}")
    print(f"{'=' * 60}")

    net_bcp = create_TMAnet()
    net_nobcp = create_TMAnet()
    model_path1 = os.path.join(test_path, 'best_model_v.pth')
    model_path2 = os.path.join(test_path, 'best_model_r.pth')

    net_bcp.load_state_dict(torch.load(str(model_path1)))
    net_nobcp.load_state_dict(torch.load(str(model_path2)))

    net_bcp.eval()
    net_nobcp.eval()
    text_embed = get_embeddings(dataset_name="KiTS19")
    avg_metric = test_all_case_average(
        net_bcp, net_nobcp, image_list, num_classes=num_classes,
        patch_size=(128, 128, 64), stride_xy=32, stride_z=32,
        metric_detail=1, nms=0,
        text_embed=text_embed, is_KiTS19=True
    )
    print(avg_metric)

    diff_uvnet_data = {
        'exp_name': exp_name,
        'dice': avg_metric[0],
        'jc': avg_metric[1],
        'hd': avg_metric[2],
        'asd': avg_metric[3],
    }
    append_to_excel(diff_uvnet_data)
    print(f"Results for {exp_name} have been saved to {EXCEL_FILENAME}")


def testLA():
    """遍历所有测试路径。"""
    for test_path in TEST_PATHS:
        try:
            test_one_path(test_path)
        except Exception as e:
            print(f"An error occurred while writing to Excel: {e}")


if __name__ == '__main__':
    cudnn.benchmark = False
    cudnn.deterministic = True
    torch.manual_seed(1345)
    torch.cuda.manual_seed(1345)
    random.seed(1345)
    np.random.seed(1345)
    init_excel_file()

    testLA()