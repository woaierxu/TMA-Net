import os
import argparse
import pathlib
import random
import time

import torch
import pdb
import torch.nn as nn
from torch.backends import cudnn
from openpyxl import Workbook, load_workbook  # 导入 openpyxl

from networks.TMANet.TMANet import TMA_Net

from networks.TMANet.embed_utils import get_embeddings
from pancreas.pancreas_utils import seed_reproducer
from pancreas.test_util import test_all_case_mean
from utils.test_3d_patch import *


# from testutildtc import *
# from test_usenet.dtc import VNet

os.environ['CUDA_VISIBLE_DEVICES'] = '0'
test_path = "model/Pancreas_TMA_fusion_lab10_per10/self_train"
exp_name = pathlib.Path(test_path).parent.name
num_classes = 2

# --- 新增: Excel 文件配置 ---
EXCEL_FILENAME = os.path.join('model/', "test_metrics_summary.xlsx")
# 定义表头，方便后续扩展
EXCEL_HEADERS = ['exp_name', 'dice', 'jc', 'hd', 'asd']
# --- 结束: 新增部分 ---


with open('../Datasets/pancreas/data_split/test.txt', 'r') as f:
    image_list = f.readlines()
image_list = ["/data1/mengqingxu/Dataset/Pancreas/preprocess-SDCL/" + item.replace('\n', '').replace('data','')
              for
              item
              in image_list]

def create_TMAnet(ema=False):
    net = TMA_Net(
        patch_size = (96,96,96),
        n_channels=1,
        n_classes=2,
        normalization='instancenorm',
        has_dropout=False,
        text_fuse_level=1
    )
    net = nn.DataParallel(net)
    model = net.cuda()
    if ema:
        for param in model.parameters():
            param.detach_()
    return model


num_classes_dfu = 2




# --- 新增: 初始化Excel文件的函数 ---
def init_excel_file():
    """检查Excel文件是否存在，如果不存在则创建并写入表头。"""
    if not os.path.exists(EXCEL_FILENAME):
        # 创建一个新的工作簿 (workbook)
        workbook = Workbook()
        # 获取活动工作表 (worksheet)
        sheet = workbook.active
        sheet.title = "Metrics"
        # 写入表头
        sheet.append(EXCEL_HEADERS)
        # 保存文件
        workbook.save(EXCEL_FILENAME)
        print(f"Created new Excel file: {EXCEL_FILENAME}")


# --- 新增: 追加数据到Excel的函数 ---
def append_to_excel(data_dict):
    """将一行数据追加到Excel文件中。"""
    try:
        # 加载已有的工作簿
        workbook = load_workbook(EXCEL_FILENAME)
        sheet = workbook.active
        # 按照表头顺序准备要写入的数据行
        row_to_append = [data_dict.get(header, "") for header in EXCEL_HEADERS]
        # 追加新行
        sheet.append(row_to_append)
        # 保存
        workbook.save(EXCEL_FILENAME)
    except FileNotFoundError:
        print(f"Error: Excel file {EXCEL_FILENAME} not found. Please initialize it first.")
    except Exception as e:
        print(f"An error occurred while writing to Excel: {e}")


def test_pancreas():
    print("exp_name: ", exp_name)
    net_bcp = create_TMAnet()
    # 接收 timesteps 的值
    net_nobcp = create_TMAnet()
    try:
        model_path1 = os.path.join(test_path, 'best_ema_10_self_v.pth')
        model_path2 = os.path.join(test_path, 'best_ema_10_self_r.pth')
        net_bcp.load_state_dict(torch.load(str(model_path1))["net"])
        net_nobcp.load_state_dict(torch.load(str(model_path2))["net"])
        print("label percent=10%")
    except:
        model_path1 = os.path.join(test_path, 'best_ema_20_self_v.pth')
        model_path2 = os.path.join(test_path, 'best_ema_20_self_r.pth')
        net_bcp.load_state_dict(torch.load(str(model_path1))["net"])
        net_nobcp.load_state_dict(torch.load(str(model_path2))["net"])
        print("label percent=20%")


    net_bcp.eval()
    net_nobcp.eval()
    text_embed = get_embeddings(dataset_name="Pancreas")
    avg_metric,_ = test_all_case_mean(net_bcp, net_nobcp, image_list, num_classes=num_classes,
                                      patch_size=(96,96,96), stride_xy=16, stride_z=16,nms=1,
                                      text_embed=text_embed)
    print(avg_metric)
    diff_uvnet_data = {
        'exp_name': pathlib.Path(test_path).parent.name,
        'dice': avg_metric[0],
        'jc': avg_metric[1],
        'hd': avg_metric[2],
        'asd': avg_metric[3],
         # 这是一个可扩展的例子
    }


    append_to_excel(diff_uvnet_data)
    print(f"Results for {exp_name} have been saved to {EXCEL_FILENAME}")



if __name__ == '__main__':
    cudnn.benchmark = False
    cudnn.deterministic = True
    torch.manual_seed(1345)
    torch.cuda.manual_seed_all(1345)
    random.seed(1345)
    np.random.seed(1345)
    seed_test = 2020
    seed_reproducer(seed=seed_test)
    # 在所有循环开始前，初始化一次Excel文件（检查并创建表头）
    init_excel_file()

    test_pancreas()


